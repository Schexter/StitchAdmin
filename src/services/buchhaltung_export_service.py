# -*- coding: utf-8 -*-
"""
BUCHHALTUNG EXPORT SERVICE
==========================
Export-Funktionen für:
- DATEV-Export (Buchungsstapel)
- ELSTER-kompatibler CSV-Export
- GoBD-konformer Export
- Excel-Export für Steuerberater

Erstellt von Hans Hahn - Alle Rechte vorbehalten
"""

import os
import csv
import json
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Any, Optional
from io import StringIO, BytesIO
import logging

logger = logging.getLogger(__name__)

# Excel-Export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DATEVExporter:
    """
    DATEV-kompatibler Export (ASCII-Format)
    
    Format: DATEV Buchungsstapel
    Kompatibel mit DATEV Unternehmen Online
    """
    
    # DATEV Header-Felder
    HEADER_FIELDS = [
        'DATEV-Format-KZ', 'Versionsnummer', 'Datenkategorie', 'Formatname',
        'Formatversion', 'Erzeugt am', 'Importiert', 'Herkunft', 'Exportiert von',
        'Importiert von', 'Berater', 'Mandant', 'WJ-Beginn', 'Sachkontenlänge',
        'Datum von', 'Datum bis', 'Bezeichnung', 'Diktatkürzel', 'Buchungstyp',
        'Rechnungslegungszweck', 'Festschreibung', 'WKZ', 'Derivatskennzeichen',
        'Kost1', 'Kost2', 'Sachkonten-Rahmen'
    ]
    
    # Buchungssatz-Felder
    BUCHUNG_FIELDS = [
        'Umsatz', 'Soll/Haben', 'WKZ Umsatz', 'Kurs', 'Basis-Umsatz', 'WKZ Basis-Umsatz',
        'Konto', 'Gegenkonto', 'BU-Schlüssel', 'Belegdatum', 'Belegfeld 1', 'Belegfeld 2',
        'Skonto', 'Buchungstext', 'Postensperre', 'Diverse Adressnummer', 'Geschäftspartnerbank',
        'Sachverhalt', 'Zinssperre', 'Beleglink', 'Beleginfo Art 1', 'Beleginfo Inhalt 1',
        'Beleginfo Art 2', 'Beleginfo Inhalt 2', 'Beleginfo Art 3', 'Beleginfo Inhalt 3',
        'Beleginfo Art 4', 'Beleginfo Inhalt 4', 'Beleginfo Art 5', 'Beleginfo Inhalt 5',
        'Beleginfo Art 6', 'Beleginfo Inhalt 6', 'Beleginfo Art 7', 'Beleginfo Inhalt 7',
        'Beleginfo Art 8', 'Beleginfo Inhalt 8', 'KOST1', 'KOST2', 'KOST-Menge',
        'EU-Land u. UStID', 'EU-Steuersatz', 'Abw. Versteuerungsart', 'Sachverhalt L+L',
        'Funktionsergänzung L+L', 'BU 49 Hauptfunktionstyp', 'BU 49 Hauptfunktionsnummer',
        'BU 49 Funktionsergänzung', 'Zusatzinformation Art 1', 'Zusatzinformation Inhalt 1',
        'Zusatzinformation Art 2', 'Zusatzinformation Inhalt 2', 'Stück', 'Gewicht',
        'Zahlweise', 'Forderungsart', 'Veranlagungsjahr', 'Zugeordnete Fälligkeit',
        'Skontotyp', 'Auftragsnummer', 'Buchungstyp', 'USt-Schlüssel', 'EU-Land',
        'Sachverhalt L+L', 'EU-Steuersatz Ursprung', 'IM EU-Steuersatz'
    ]
    
    def __init__(self, berater_nr: str = '0', mandant_nr: str = '0', 
                 wj_beginn: date = None, kontenrahmen: str = 'SKR03'):
        """
        Initialisiert DATEV-Exporter
        
        Args:
            berater_nr: DATEV Beraternummer
            mandant_nr: DATEV Mandantennummer
            wj_beginn: Beginn des Wirtschaftsjahres
            kontenrahmen: SKR03 oder SKR04
        """
        self.berater_nr = berater_nr
        self.mandant_nr = mandant_nr
        self.wj_beginn = wj_beginn or date(date.today().year, 1, 1)
        self.kontenrahmen = kontenrahmen
    
    def export_buchungen(self, buchungen: List, datum_von: date, datum_bis: date) -> str:
        """
        Exportiert Buchungen im DATEV-Format
        
        Args:
            buchungen: Liste von Buchung-Objekten
            datum_von: Startdatum
            datum_bis: Enddatum
        
        Returns:
            CSV-String im DATEV-Format
        """
        output = StringIO()
        writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        
        # Header-Zeile
        header = self._create_header(datum_von, datum_bis)
        writer.writerow(header)
        
        # Spaltenüberschriften
        writer.writerow(self.BUCHUNG_FIELDS[:20])  # Wichtigste Felder
        
        # Buchungen
        for buchung in buchungen:
            row = self._format_buchung(buchung)
            writer.writerow(row)
        
        return output.getvalue()
    
    def _create_header(self, datum_von: date, datum_bis: date) -> List[str]:
        """Erstellt DATEV-Header"""
        return [
            'EXTF',  # DATEV-Format
            '700',   # Version
            '21',    # Buchungsstapel
            'Buchungsstapel',
            '12',    # Format-Version
            datetime.now().strftime('%Y%m%d%H%M%S000'),
            '',      # Importiert
            'RE',    # Herkunft (Rechnungswesen)
            'StitchAdmin',
            '',
            self.berater_nr,
            self.mandant_nr,
            self.wj_beginn.strftime('%Y%m%d'),
            '4',     # Sachkontenlänge
            datum_von.strftime('%Y%m%d'),
            datum_bis.strftime('%Y%m%d'),
            f'Export {datum_von.strftime("%m/%Y")}',
            '',      # Diktatkürzel
            '1',     # Buchungstyp
            '',      # Rechnungslegungszweck
            '0',     # Festschreibung
            'EUR',   # Währung
            '',      # Derivat
            '',      # Kost1
            '',      # Kost2
            self.kontenrahmen
        ]
    
    def _format_buchung(self, buchung) -> List[str]:
        """Formatiert eine Buchung für DATEV"""
        # Betrag formatieren (DATEV: Komma als Dezimaltrenner, kein Tausender)
        betrag = str(abs(buchung.betrag_brutto or buchung.betrag_netto)).replace('.', ',')
        
        # Soll/Haben
        if buchung.buchungs_art == 'einnahme':
            soll_haben = 'H'  # Haben (Ertrag)
        else:
            soll_haben = 'S'  # Soll (Aufwand)
        
        # Konten
        konto = buchung.konto.kontonummer if buchung.konto else ''
        gegenkonto = buchung.gegenkonto.kontonummer if buchung.gegenkonto else ''
        
        # Datum (DDMM)
        belegdatum = buchung.buchungsdatum.strftime('%d%m') if buchung.buchungsdatum else ''
        
        return [
            betrag,
            soll_haben,
            'EUR',   # Währung
            '',      # Kurs
            '',      # Basis-Umsatz
            '',      # WKZ Basis
            konto,
            gegenkonto,
            '',      # BU-Schlüssel
            belegdatum,
            buchung.belegnummer or '',
            '',      # Belegfeld 2
            '',      # Skonto
            buchung.buchungstext[:60] if buchung.buchungstext else '',  # Max 60 Zeichen
            '',      # Postensperre
            '',      # Adressnummer
            '',      # Bank
            '',      # Sachverhalt
            '',      # Zinssperre
            ''       # Beleglink
        ]


class ELSTERExporter:
    """
    ELSTER-kompatibler CSV-Export für USt-Voranmeldung
    
    Erstellt CSV-Datei die in ELSTER-Portal importiert werden kann
    oder als Vorlage für manuelle Eingabe dient
    """
    
    # ELSTER Kennzahlen für USt-VA
    KENNZAHLEN = {
        'kz81': 'Steuerpflichtige Umsätze 19%',
        'kz86': 'Steuerpflichtige Umsätze 7%',
        'kz35': 'Steuerfreie Umsätze mit Vorsteuerabzug',
        'kz77': 'Innergemeinschaftliche Lieferungen',
        'kz76': 'Umsätze für die der Leistungsempfänger Steuer schuldet',
        'kz66': 'Vorsteuer aus Rechnungen',
        'kz61': 'Vorsteuer aus innergemeinschaftlichem Erwerb',
        'kz83': 'Verbleibende USt-Vorauszahlung'
    }
    
    def export_ust_voranmeldung(self, voranmeldung) -> str:
        """
        Exportiert USt-Voranmeldung als CSV
        
        Args:
            voranmeldung: UStVoranmeldung-Objekt
        
        Returns:
            CSV-String
        """
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Header
        writer.writerow(['ELSTER USt-Voranmeldung Export'])
        writer.writerow(['Erstellt am:', datetime.now().strftime('%d.%m.%Y %H:%M')])
        writer.writerow(['Zeitraum:', f'{voranmeldung.zeitraum_von} bis {voranmeldung.zeitraum_bis}'])
        writer.writerow([])
        
        # Kennzahlen
        writer.writerow(['Kennzahl', 'Bezeichnung', 'Bemessungsgrundlage', 'Steuer'])
        
        # Umsätze
        writer.writerow(['81', 'Umsätze 19%', self._fmt(voranmeldung.umsatz_19_netto), self._fmt(voranmeldung.ust_19)])
        writer.writerow(['86', 'Umsätze 7%', self._fmt(voranmeldung.umsatz_7_netto), self._fmt(voranmeldung.ust_7)])
        
        # Innergemeinschaftliche Erwerbe
        if voranmeldung.ig_erwerbe_netto:
            writer.writerow(['89', 'IG Erwerbe', self._fmt(voranmeldung.ig_erwerbe_netto), self._fmt(voranmeldung.ust_ig_erwerbe)])
        
        writer.writerow([])
        
        # Vorsteuer
        writer.writerow(['Vorsteuer', '', '', ''])
        writer.writerow(['66', 'Vorsteuer 19%', '', self._fmt(voranmeldung.vorsteuer_19)])
        writer.writerow(['66', 'Vorsteuer 7%', '', self._fmt(voranmeldung.vorsteuer_7)])
        writer.writerow(['61', 'Vorsteuer IG', '', self._fmt(voranmeldung.vorsteuer_ig)])
        writer.writerow(['', 'Vorsteuer gesamt', '', self._fmt(voranmeldung.vorsteuer_gesamt)])
        
        writer.writerow([])
        
        # Ergebnis
        writer.writerow(['83', 'Verbleibende USt-Vorauszahlung', '', self._fmt(voranmeldung.ust_zahllast)])
        
        if voranmeldung.ust_zahllast > 0:
            writer.writerow(['', 'ZAHLUNG AN FINANZAMT', '', self._fmt(voranmeldung.ust_zahllast)])
        else:
            writer.writerow(['', 'ERSTATTUNG VOM FINANZAMT', '', self._fmt(abs(voranmeldung.ust_zahllast))])
        
        return output.getvalue()
    
    def _fmt(self, value) -> str:
        """Formatiert Dezimalwert für Export"""
        if value is None:
            return '0,00'
        return str(value).replace('.', ',')


class GoBDExporter:
    """
    GoBD-konformer Export (Grundsätze ordnungsmäßiger Buchführung)
    
    Erstellt vollständige, unveränderbare Export-Pakete mit:
    - Buchungsjournal
    - Stammdaten
    - Verfahrensdokumentation
    - Prüfsummen
    """
    
    def export_gobd_paket(self, jahr: int, buchungen: List, 
                          kunden: List, lieferanten: List) -> Dict[str, Any]:
        """
        Erstellt GoBD-konformes Exportpaket
        
        Args:
            jahr: Geschäftsjahr
            buchungen: Alle Buchungen des Jahres
            kunden: Kundenstammdaten
            lieferanten: Lieferantenstammdaten
        
        Returns:
            Dict mit Dateiname -> Inhalt
        """
        paket = {}
        
        # 1. Buchungsjournal
        paket['buchungen.csv'] = self._export_buchungen_csv(buchungen)
        
        # 2. Kundenstammdaten
        paket['kunden.csv'] = self._export_kunden_csv(kunden)
        
        # 3. Lieferantenstammdaten
        paket['lieferanten.csv'] = self._export_lieferanten_csv(lieferanten)
        
        # 4. Index-Datei (gdpdu-01-09-2004.xml kompatibel)
        paket['index.xml'] = self._create_index_xml(jahr)
        
        # 5. Prüfsummen
        paket['checksums.txt'] = self._create_checksums(paket)
        
        return paket
    
    def _export_buchungen_csv(self, buchungen: List) -> str:
        """Exportiert Buchungen als CSV"""
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Header
        writer.writerow([
            'Buchungsnummer', 'Buchungsdatum', 'Erfassungsdatum', 
            'Belegnummer', 'Buchungstext', 'Soll-Konto', 'Haben-Konto',
            'Betrag Netto', 'MwSt-Satz', 'MwSt-Betrag', 'Betrag Brutto',
            'Kunde/Lieferant', 'Kostenstelle'
        ])
        
        for b in buchungen:
            writer.writerow([
                b.id,
                b.buchungsdatum.strftime('%d.%m.%Y') if b.buchungsdatum else '',
                b.erfassungsdatum.strftime('%d.%m.%Y %H:%M') if b.erfassungsdatum else '',
                b.belegnummer or '',
                b.buchungstext or '',
                b.soll_konto.kontonummer if b.soll_konto else '',
                b.haben_konto.kontonummer if b.haben_konto else '',
                str(b.betrag_netto or 0).replace('.', ','),
                str(b.mwst_satz or 0).replace('.', ','),
                str(b.mwst_betrag or 0).replace('.', ','),
                str(b.betrag_brutto or 0).replace('.', ','),
                b.kunde.display_name if b.kunde else '',
                b.kostenstelle.nummer if b.kostenstelle else ''
            ])
        
        return output.getvalue()
    
    def _export_kunden_csv(self, kunden: List) -> str:
        """Exportiert Kundenstammdaten"""
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        writer.writerow([
            'Kundennummer', 'Firma', 'Vorname', 'Nachname', 
            'Straße', 'PLZ', 'Ort', 'Land', 'USt-IdNr', 'Angelegt am'
        ])
        
        for k in kunden:
            writer.writerow([
                k.customer_number or k.id,
                k.company_name or '',
                k.first_name or '',
                k.last_name or '',
                k.street or '',
                k.postal_code or '',
                k.city or '',
                k.country or 'DE',
                k.vat_id or '',
                k.created_at.strftime('%d.%m.%Y') if k.created_at else ''
            ])
        
        return output.getvalue()
    
    def _export_lieferanten_csv(self, lieferanten: List) -> str:
        """Exportiert Lieferantenstammdaten"""
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        writer.writerow([
            'Lieferantennummer', 'Firma', 'Straße', 'PLZ', 'Ort', 
            'USt-IdNr', 'IBAN', 'BIC'
        ])
        
        for l in lieferanten:
            writer.writerow([
                l.id,
                l.name or '',
                l.street or '',
                l.postal_code or '',
                l.city or '',
                l.tax_id or '',
                l.iban or '',
                l.bic or ''
            ])
        
        return output.getvalue()
    
    def _create_index_xml(self, jahr: int) -> str:
        """Erstellt GDPdU-Index (vereinfacht)"""
        return f'''<?xml version="1.0" encoding="UTF-8"?>
<DataSet xmlns="urn:gdpdu:2004-09-01" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
    <Version>1.0</Version>
    <DataSupplier>
        <Name>StitchAdmin 2.0</Name>
        <Location>Deutschland</Location>
        <Comment>GoBD-Export {jahr}</Comment>
    </DataSupplier>
    <Media>
        <Name>Buchungsdaten {jahr}</Name>
        <Table>
            <Name>Buchungsjournal</Name>
            <URL>buchungen.csv</URL>
            <DecimalSymbol>,</DecimalSymbol>
            <DigitGroupingSymbol>.</DigitGroupingSymbol>
            <VariableLength>
                <ColumnDelimiter>;</ColumnDelimiter>
                <RecordDelimiter>&#10;</RecordDelimiter>
                <TextEncapsulator>"</TextEncapsulator>
            </VariableLength>
        </Table>
        <Table>
            <Name>Kunden</Name>
            <URL>kunden.csv</URL>
        </Table>
        <Table>
            <Name>Lieferanten</Name>
            <URL>lieferanten.csv</URL>
        </Table>
    </Media>
</DataSet>'''
    
    def _create_checksums(self, paket: Dict) -> str:
        """Erstellt Prüfsummen für alle Dateien"""
        import hashlib
        
        lines = ['GoBD Export Prüfsummen', f'Erstellt: {datetime.now().isoformat()}', '']
        
        for filename, content in paket.items():
            if filename != 'checksums.txt':
                md5 = hashlib.md5(content.encode('utf-8')).hexdigest()
                sha256 = hashlib.sha256(content.encode('utf-8')).hexdigest()
                lines.append(f'{filename}:')
                lines.append(f'  MD5: {md5}')
                lines.append(f'  SHA256: {sha256}')
                lines.append('')
        
        return '\n'.join(lines)


class ExcelExporter:
    """
    Excel-Export für Steuerberater und Auswertungen
    """
    
    def export_buchungsjournal(self, buchungen: List, titel: str = 'Buchungsjournal') -> bytes:
        """Exportiert Buchungen als Excel"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl nicht installiert!")
        
        wb = Workbook()
        ws = wb.active
        ws.title = titel[:31]  # Excel max 31 Zeichen
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        headers = ['Nr.', 'Datum', 'Beleg', 'Buchungstext', 'Soll', 'Haben', 
                   'Netto', 'MwSt %', 'MwSt', 'Brutto', 'Kunde/Lieferant']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Daten
        for row_idx, b in enumerate(buchungen, 2):
            ws.cell(row=row_idx, column=1, value=b.id)
            ws.cell(row=row_idx, column=2, value=b.buchungsdatum)
            ws.cell(row=row_idx, column=3, value=b.belegnummer)
            ws.cell(row=row_idx, column=4, value=b.buchungstext)
            ws.cell(row=row_idx, column=5, value=b.soll_konto.kontonummer if b.soll_konto else '')
            ws.cell(row=row_idx, column=6, value=b.haben_konto.kontonummer if b.haben_konto else '')
            ws.cell(row=row_idx, column=7, value=float(b.betrag_netto or 0))
            ws.cell(row=row_idx, column=8, value=float(b.mwst_satz or 0))
            ws.cell(row=row_idx, column=9, value=float(b.mwst_betrag or 0))
            ws.cell(row=row_idx, column=10, value=float(b.betrag_brutto or 0))
            ws.cell(row=row_idx, column=11, value=b.kunde.display_name if b.kunde else '')
            
            # Währungsformat
            for col in [7, 9, 10]:
                ws.cell(row=row_idx, column=col).number_format = '#,##0.00 €'
        
        # Spaltenbreiten
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 25
        
        # In Bytes konvertieren
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_bwa(self, bwa_daten: Dict) -> bytes:
        """Exportiert BWA als Excel"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl nicht installiert!")
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'BWA'
        
        # Styles
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11)
        currency_format = '#,##0.00 €'
        
        row = 1
        
        # Titel
        ws.cell(row=row, column=1, value='Betriebswirtschaftliche Auswertung')
        ws.cell(row=row, column=1).font = Font(bold=True, size=16)
        row += 1
        
        ws.cell(row=row, column=1, value=f"Zeitraum: {bwa_daten.get('zeitraum', '')}")
        row += 2
        
        # Erlöse
        ws.cell(row=row, column=1, value='ERLÖSE')
        ws.cell(row=row, column=1).font = section_font
        row += 1
        
        for konto, betrag in bwa_daten.get('erloese', {}).items():
            ws.cell(row=row, column=1, value=konto)
            ws.cell(row=row, column=2, value=float(betrag))
            ws.cell(row=row, column=2).number_format = currency_format
            row += 1
        
        ws.cell(row=row, column=1, value='Summe Erlöse')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(bwa_daten.get('summe_erloese', 0)))
        ws.cell(row=row, column=2).number_format = currency_format
        ws.cell(row=row, column=2).font = Font(bold=True)
        row += 2
        
        # Aufwendungen
        ws.cell(row=row, column=1, value='AUFWENDUNGEN')
        ws.cell(row=row, column=1).font = section_font
        row += 1
        
        for konto, betrag in bwa_daten.get('aufwendungen', {}).items():
            ws.cell(row=row, column=1, value=konto)
            ws.cell(row=row, column=2, value=float(betrag))
            ws.cell(row=row, column=2).number_format = currency_format
            row += 1
        
        ws.cell(row=row, column=1, value='Summe Aufwendungen')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(bwa_daten.get('summe_aufwendungen', 0)))
        ws.cell(row=row, column=2).number_format = currency_format
        ws.cell(row=row, column=2).font = Font(bold=True)
        row += 2
        
        # Ergebnis
        ws.cell(row=row, column=1, value='BETRIEBSERGEBNIS')
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value=float(bwa_daten.get('betriebsergebnis', 0)))
        ws.cell(row=row, column=2).number_format = currency_format
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        
        # Spaltenbreiten
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
